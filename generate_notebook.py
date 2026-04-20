# =============================================================================
# generate_notebook.py
# Chay file nay 1 lan de tao spam_classifier.ipynb
# Lenh: python generate_notebook.py
# =============================================================================

import sys
sys.stdout.reconfigure(encoding="utf-8")
import nbformat as nbf
import os

nb = nbf.v4.new_notebook()
nb.metadata = {
    "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
    "language_info": {"name": "python", "version": "3.13.0"}
}

cells = []

def md(src):
    return nbf.v4.new_markdown_cell(src)

def code(src):
    return nbf.v4.new_code_cell(src)

# =============================================================================
# CELL 1 — TIÊU ĐỀ
# =============================================================================
cells.append(md("""# 🔍 Phân Loại Email Rác — Spam Email Classifier
## Chủ đề 16 | Trí Tuệ Nhân Tạo | UTH

| Thông tin | Chi tiết |
|-----------|----------|
| **Mô hình** | Multinomial Naive Bayes |
| **Dữ liệu** | 2 tài khoản Gmail thực (~39,000 email) |
| **Ngôn ngữ** | Tiếng Việt + Tiếng Anh |
| **Công nghệ** | Python · scikit-learn · underthesea · TF-IDF |

---
**Hướng dẫn chạy:** Chọn `Run All` hoặc chạy từng cell theo thứ tự từ trên xuống.
Lần đầu mất ~3-5 phút (đọc mbox + tiền xử lý). Các lần sau chạy nhanh hơn nhờ cache.
"""))

# =============================================================================
# CELL 2 — IMPORT THƯ VIỆN
# =============================================================================
cells.append(code("""\
# ============================================================
# IMPORT TẤT CẢ THƯ VIỆN
# ============================================================
import sys
import os
import re
import pickle
import warnings
import mailbox
import email.header
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
import seaborn as sns
from tqdm.notebook import tqdm          # thanh tiến trình cho Jupyter
from collections import Counter
from IPython.display import display, HTML

# --- Scikit-learn ---
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.naive_bayes import MultinomialNB
from sklearn.model_selection import train_test_split
from sklearn.metrics import (
    classification_report, confusion_matrix,
    accuracy_score, precision_score, recall_score, f1_score
)

# --- Xử lý mất cân bằng ---
try:
    from imblearn.over_sampling import RandomOverSampler
    IMBLEARN_OK = True
except ImportError:
    IMBLEARN_OK = False
    print("⚠️  imbalanced-learn chưa cài → pip install imbalanced-learn")

# --- Tách từ tiếng Việt ---
try:
    from underthesea import word_tokenize as vn_tokenize
    UNDERTHESEA_OK = True
except ImportError:
    UNDERTHESEA_OK = False
    print("⚠️  underthesea chưa cài → pip install underthesea")

# --- Báo cáo ---
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from tabulate import tabulate

warnings.filterwarnings('ignore')

# Thiết lập đồ thị — dùng Arial để hiển thị tiếng Việt trên Windows
matplotlib.rcParams['font.family'] = 'DejaVu Sans'
matplotlib.rcParams['figure.dpi']  = 100
plt.style.use('seaborn-v0_8-whitegrid')

print("✅ Tất cả thư viện đã được import thành công!")
print(f"   Python  : {sys.version.split()[0]}")
print(f"   Pandas  : {pd.__version__}")
print(f"   Sklearn : {__import__('sklearn').__version__}")
"""))

# =============================================================================
# CELL 3 — CẤU HÌNH DỰ ÁN
# =============================================================================
cells.append(md("## ⚙️ Cấu hình Dự án"))

cells.append(code("""\
# ============================================================
# CẤU HÌNH — chỉnh sửa đường dẫn nếu cần
# ============================================================

# Thư mục gốc của dự án (tự động lấy vị trí notebook)
BASE_DIR = os.getcwd()

# ⚠️  THAY ĐỔI đường dẫn này nếu bạn đặt file mbox ở nơi khác
MBOX_PATH = r"D:\\UTH\\AI\\Spam\\data\\Mail\\Tất cả thư bao gồm spam và thư rác.mbox"

# File CSV dữ liệu thô
MAIL1_CSV    = os.path.join(BASE_DIR, "mail1.csv")     # Gmail 1
MESSAGES_CSV = os.path.join(BASE_DIR, "messages.csv")  # Gmail 2 (dự phòng)

# Thư mục đầu ra (tự động tạo)
OUTPUT_DIR  = os.path.join(BASE_DIR, "output")
REPORTS_DIR = os.path.join(OUTPUT_DIR, "reports")
MODELS_DIR  = os.path.join(OUTPUT_DIR, "models")
for d in [OUTPUT_DIR, REPORTS_DIR, MODELS_DIR]:
    os.makedirs(d, exist_ok=True)

# File cache & kết quả
LABELED_CSV   = os.path.join(OUTPUT_DIR, "labeled_dataset.csv")
PROCESSED_CSV = os.path.join(OUTPUT_DIR, "processed_dataset.csv")
MODEL_PKL     = os.path.join(MODELS_DIR, "naive_bayes_model.pkl")
TFIDF_PKL     = os.path.join(MODELS_DIR, "tfidf_vectorizer.pkl")
REPORT_XLSX   = os.path.join(REPORTS_DIR, "spam_report.xlsx")
REPORT_TXT    = os.path.join(REPORTS_DIR, "spam_summary.txt")

# Nhãn
SPAM, HAM = "spam", "ham"

# Tham số mô hình
TEST_SIZE    = 0.20   # 20% dữ liệu dùng để test
RANDOM_STATE = 42     # seed cố định để kết quả tái lặp
MAX_FEATURES = 50000  # số từ tối đa trong TF-IDF
NGRAM_RANGE  = (1, 2) # dùng cả unigram (1 từ) và bigram (2 từ liên tiếp)

# Nhãn Gmail2 được tính là SPAM
GMAIL_SPAM_LABELS = ["Spam", "Danh mục Khuyến mại"]

# Heuristics cho Gmail1 (không có mbox)
SPAM_SENDERS = [
    "ecomm.lenovo.com", "recommendations@ted.com", "recommends@ted.com",
    "discover.pinterest.com", "inspire.pinterest.com", "ideas.pinterest.com",
    "explore.pinterest.com", "pinterest.com", "no-reply@grab.com",
    "facebookmail.com", "news@insideapple.apple.com", "noreply@autocode.com",
]
HAM_SENDERS = [
    "security@facebookmail.com", "accounts.google.com",
    "forms-receipts-noreply@google.com", "appstore@insideapple.apple.com",
    "no_reply@email.apple.com",
]
SPAM_KEYWORDS = [
    "ưu đãi", "khuyến mãi", "giảm giá", "voucher", "deal", "offer",
    "sale", "promo", "discount", "unsubscribe", "miễn phí", "free",
    "recommendations", "recommended for you",
]
PHISHING_KW = [
    "tài khoản của bạn bị", "your account has been", "password reset",
    "verify your account", "xác minh tài khoản", "trúng thưởng", "bạn đã thắng",
]

print("✅ Cấu hình đã thiết lập!")
print(f"   📁 Thư mục dự án : {BASE_DIR}")
print(f"   📂 Thư mục đầu ra: {OUTPUT_DIR}")
print(f"   📄 File mbox     : {'✅ Tìm thấy' if os.path.exists(MBOX_PATH) else '⚠️  Không thấy (sẽ dùng CSV fallback)'}")
"""))

# =============================================================================
# CELL 4 — HÀM GÁN NHÃN
# =============================================================================
cells.append(md("## 📂 Phase 1 — Thu thập & Gán nhãn Dữ liệu"))

cells.append(code("""\
# ============================================================
# CÁC HÀM GÁN NHÃN DỮ LIỆU
# ============================================================

def decode_header(raw):
    \"\"\"Giải mã header email bị mã hóa RFC2047 (=?UTF-8?B?...?=).\"\"\"
    if not raw:
        return ""
    try:
        parts = email.header.decode_header(raw)
        result = ""
        for part, enc in parts:
            if isinstance(part, bytes):
                result += part.decode(enc or "utf-8", errors="replace")
            else:
                result += str(part)
        return result.strip()
    except Exception:
        return str(raw).strip()


def extract_body(msg):
    \"\"\"Trích xuất nội dung text/plain từ đối tượng email mailbox.\"\"\"
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            if "attachment" in str(part.get("Content-Disposition", "")):
                continue
            if ctype == "text/plain":
                raw = part.get_payload(decode=True)
                if raw:
                    charset = part.get_content_charset() or "utf-8"
                    body = raw.decode(charset, errors="replace")
                    break
            elif ctype == "text/html" and not body:
                raw = part.get_payload(decode=True)
                if raw:
                    charset = part.get_content_charset() or "utf-8"
                    body = raw.decode(charset, errors="replace")
    else:
        raw = msg.get_payload(decode=True)
        if raw:
            charset = msg.get_content_charset() or "utf-8"
            body = raw.decode(charset, errors="replace")
    return body[:8000]  # giới hạn 8000 ký tự để tiết kiệm bộ nhớ


def label_by_gmail(label_str):
    \"\"\"Phân loại spam/ham từ nhãn Gmail thật (X-Gmail-Labels).\"\"\"
    if not label_str:
        return HAM
    labels = [l.strip() for l in label_str.split(",")]
    return SPAM if any(sl in labels for sl in GMAIL_SPAM_LABELS) else HAM


def label_by_heuristics(sender, subject):
    \"\"\"Phân loại spam/ham bằng quy tắc (cho Gmail1 không có mbox).\"\"\"
    s = sender.lower()
    q = subject.lower()
    # Ưu tiên 1: sender quan trọng → chắc chắn ham
    if any(h in s for h in HAM_SENDERS):
        return HAM
    # Ưu tiên 2: sender marketing → spam
    if any(sp in s for sp in SPAM_SENDERS):
        return SPAM
    # Ưu tiên 3: từ khóa spam trong tiêu đề
    if any(kw in q for kw in SPAM_KEYWORDS):
        return SPAM
    return HAM


print("✅ Hàm gán nhãn sẵn sàng!")
"""))

# =============================================================================
# CELL 5 — TẢI DỮ LIỆU (CÓ CACHE)
# =============================================================================
cells.append(code("""\
# ============================================================
# TẢI VÀ GÁN NHÃN DỮ LIỆU (CÓ CACHE)
# Lần đầu: xử lý từ mbox/CSV → lưu cache (~3-5 phút)
# Lần sau: tải từ cache (~5 giây)
# ============================================================

def load_gmail2():
    \"\"\"Đọc Gmail2 từ file .mbox — lấy nhãn thật X-Gmail-Labels.\"\"\"
    records = []
    try:
        mbox = mailbox.mbox(MBOX_PATH)
        for msg in tqdm(mbox, desc="📬 Gmail 2 (mbox)", unit="email"):
            try:
                label_str = decode_header(msg.get("X-Gmail-Labels", ""))
                label     = label_by_gmail(label_str)
                subject   = decode_header(msg.get("Subject", ""))
                sender    = decode_header(msg.get("From", ""))
                body      = extract_body(msg)
                records.append({
                    "source": "gmail2", "subject": subject, "sender": sender,
                    "date": msg.get("Date", ""), "body": body,
                    "label": label, "label_source": "gmail_labels"
                })
            except Exception:
                continue
        mbox.close()
    except FileNotFoundError:
        print("   ⚠️  Không tìm thấy mbox → dùng messages.csv (fallback)")
        df2 = pd.read_csv(MESSAGES_CSV, encoding="utf-8-sig",
                          header=0, low_memory=False)
        df2.columns = ["subject","sender","to","date","starred","size","body"]
        df2.fillna("", inplace=True)
        for _, row in tqdm(df2.iterrows(), total=len(df2),
                           desc="📬 Gmail 2 (CSV fallback)", unit="email"):
            label = label_by_heuristics(str(row["sender"]), str(row["subject"]))
            records.append({
                "source": "gmail2", "subject": str(row["subject"]),
                "sender": str(row["sender"]), "date": str(row["date"]),
                "body": str(row["body"]), "label": label,
                "label_source": "heuristic"
            })
    s = sum(1 for r in records if r["label"] == SPAM)
    print(f"   ✅ Gmail 2: {len(records):,} email → {s:,} spam | {len(records)-s:,} ham")
    return records


def load_gmail1():
    \"\"\"Đọc Gmail1 từ mail1.csv — gán nhãn bằng heuristics.\"\"\"
    records = []
    df1 = pd.read_csv(MAIL1_CSV, encoding="utf-8-sig",
                      header=0, low_memory=False)
    df1.columns = ["subject","sender","to","date","starred","size","body"]
    df1.fillna("", inplace=True)
    for _, row in tqdm(df1.iterrows(), total=len(df1),
                       desc="📬 Gmail 1 (heuristics)", unit="email"):
        label = label_by_heuristics(str(row["sender"]), str(row["subject"]))
        records.append({
            "source": "gmail1", "subject": str(row["subject"]),
            "sender": str(row["sender"]), "date": str(row["date"]),
            "body": str(row["body"]), "label": label,
            "label_source": "heuristic"
        })
    s = sum(1 for r in records if r["label"] == SPAM)
    print(f"   ✅ Gmail 1: {len(records):,} email → {s:,} spam | {len(records)-s:,} ham")
    return records


# Kiểm tra cache
if os.path.exists(LABELED_CSV):
    print(f"📦 Tìm thấy cache: {LABELED_CSV}")
    df = pd.read_csv(LABELED_CSV, encoding="utf-8-sig", low_memory=False)
    df.fillna("", inplace=True)
    print(f"   ✅ Tải cache thành công: {len(df):,} email")
else:
    print("🔄 Chưa có cache — xử lý từ đầu...")
    recs = load_gmail2() + load_gmail1()
    df   = pd.DataFrame(recs)

    # Sửa nhãn phishing (bị gán nhầm ham)
    fixed = 0
    for idx, row in df.iterrows():
        if row["label"] == HAM:
            combo = (str(row["subject"]) + " " + str(row["body"])[:500]).lower()
            if any(kw in combo for kw in PHISHING_KW):
                df.at[idx, "label"] = SPAM
                fixed += 1
    print(f"   🔧 Đã sửa {fixed} email phishing → spam")

    df.to_csv(LABELED_CSV, index=False, encoding="utf-8-sig")
    print(f"   💾 Đã lưu cache: {LABELED_CSV}")

# Reset index để đảm bảo index liên tục 0,1,2,...
df = df.reset_index(drop=True)

spam_n = (df["label"] == SPAM).sum()
ham_n  = (df["label"] == HAM).sum()
print(f"\\n📊 TỔNG KẾT DỮ LIỆU:")
print(f"   Tổng  : {len(df):,} email")
print(f"   🔴 Spam: {spam_n:,} ({spam_n/len(df)*100:.1f}%)")
print(f"   🟢 Ham : {ham_n:,}  ({ham_n/len(df)*100:.1f}%)")
"""))

# =============================================================================
# CELL 6 — EDA
# =============================================================================
cells.append(md("## 📊 Phase 2 — Phân tích Khám phá Dữ liệu (EDA)"))

cells.append(code("""\
# ============================================================
# EDA — 4 BIỂU ĐỒ PHÂN TÍCH
# ============================================================

fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle("📊 Phân tích Dữ liệu Email — EDA",
             fontsize=16, fontweight='bold', y=1.01)

SPAM_COLOR = '#e74c3c'   # đỏ = spam
HAM_COLOR  = '#2ecc71'   # xanh = ham

# --- Biểu đồ 1: Bar chart số lượng nhãn ---
counts = df["label"].value_counts()
s_cnt  = counts.get(SPAM, 0)
h_cnt  = counts.get(HAM,  0)
bars = axes[0,0].bar(['SPAM', 'HAM'], [s_cnt, h_cnt],
                     color=[SPAM_COLOR, HAM_COLOR],
                     edgecolor='white', linewidth=1.5, width=0.5)
axes[0,0].set_title('Phân phối nhãn', fontweight='bold')
axes[0,0].set_ylabel('Số lượng email')
for bar in bars:
    axes[0,0].text(bar.get_x() + bar.get_width()/2,
                   bar.get_height() + max(s_cnt,h_cnt)*0.01,
                   f'{int(bar.get_height()):,}',
                   ha='center', va='bottom', fontweight='bold')
axes[0,0].grid(axis='y', alpha=0.3)

# --- Biểu đồ 2: Pie chart tỉ lệ ---
s_pct = s_cnt/len(df)*100
h_pct = h_cnt/len(df)*100
wedges, texts, autotexts = axes[0,1].pie(
    [s_cnt, h_cnt],
    labels=[f'SPAM\\n{s_pct:.1f}%', f'HAM\\n{h_pct:.1f}%'],
    colors=[SPAM_COLOR, HAM_COLOR],
    autopct='%1.1f%%', startangle=90,
    wedgeprops={'edgecolor': 'white', 'linewidth': 2}
)
for at in autotexts:
    at.set_fontweight('bold')
axes[0,1].set_title('Tỉ lệ Spam vs Ham', fontweight='bold')

# --- Biểu đồ 3: Phân phối độ dài tiêu đề ---
df['subj_len'] = df['subject'].astype(str).str.len()
axes[1,0].hist(df[df['label']==SPAM]['subj_len'].clip(0, 150),
               bins=35, alpha=0.65, color=SPAM_COLOR, label='Spam', density=True)
axes[1,0].hist(df[df['label']==HAM]['subj_len'].clip(0, 150),
               bins=35, alpha=0.65, color=HAM_COLOR,  label='Ham',  density=True)
axes[1,0].set_title('Phân phối độ dài tiêu đề', fontweight='bold')
axes[1,0].set_xlabel('Số ký tự trong tiêu đề')
axes[1,0].set_ylabel('Mật độ')
axes[1,0].legend()
axes[1,0].grid(alpha=0.3)

# --- Biểu đồ 4: Top 10 domain người gửi ---
def get_domain(sender):
    m = re.search(r'@([\\w.\\-]+)', str(sender))
    return m.group(1).lower() if m else 'unknown'

df['domain'] = df['sender'].apply(get_domain)
top_domains  = df['domain'].value_counts().head(10)
axes[1,1].barh(range(len(top_domains)),
               top_domains.values[::-1] if len(top_domains) > 0 else [],
               color='#3498db', alpha=0.8)
axes[1,1].set_yticks(range(len(top_domains)))
axes[1,1].set_yticklabels(top_domains.index[::-1], fontsize=9)
axes[1,1].set_title('Top 10 Domain gửi email', fontweight='bold')
axes[1,1].set_xlabel('Số email')
axes[1,1].grid(axis='x', alpha=0.3)

plt.tight_layout()
plt.savefig(os.path.join(REPORTS_DIR, 'eda_analysis.png'),
            dpi=150, bbox_inches='tight', facecolor='white')
plt.show()
print("📈 Biểu đồ EDA đã lưu: output/reports/eda_analysis.png")
"""))

# =============================================================================
# CELL 7 — BẢNG THỐNG KÊ
# =============================================================================
cells.append(code("""\
# Bảng thống kê chi tiết
g1 = (df['source'] == 'gmail1').sum() if 'source' in df.columns else 0
g2 = (df['source'] == 'gmail2').sum() if 'source' in df.columns else 0

stats = [
    ["Tổng số email",           f"{len(df):,}",         ""],
    ["Email SPAM",               f"{spam_n:,}",          f"{spam_n/len(df)*100:.1f}%"],
    ["Email HAM",                f"{ham_n:,}",           f"{ham_n/len(df)*100:.1f}%"],
    ["Gmail 1 (wiisch3009)",     f"{g1:,}",              "Gán nhãn heuristics"],
    ["Gmail 2 (nguyenkhanhtoan)",f"{g2:,}",              "Nhãn thật từ Gmail"],
    ["Độ dài tiêu đề TB (Spam)", f"{df[df['label']==SPAM]['subj_len'].mean():.0f} ký tự", ""],
    ["Độ dài tiêu đề TB (Ham)",  f"{df[df['label']==HAM]['subj_len'].mean():.0f} ký tự",  ""],
]
print(tabulate(stats, headers=["Chỉ số", "Giá trị", "Ghi chú"], tablefmt="grid"))
"""))

# =============================================================================
# CELL 8 — TIỀN XỬ LÝ
# =============================================================================
cells.append(md("## 🔧 Phase 3 — Tiền xử lý Văn bản (Song ngữ Việt + Anh)"))

cells.append(code("""\
# ============================================================
# CÁC HÀM TIỀN XỬ LÝ VĂN BẢN
# ============================================================
from sklearn.feature_extraction.text import ENGLISH_STOP_WORDS

# Stopwords tiếng Việt
VN_STOP = {
    "và","của","cho","là","có","được","trong","không","với","các","một",
    "để","này","từ","tôi","bạn","đã","đang","sẽ","khi","như","hay","hoặc",
    "nhưng","mà","thì","vì","bởi","nên","ra","đi","lên","xuống","theo",
    "về","qua","trên","dưới","sau","trước","đây","đó","ở","tại","vào",
    "rất","lắm","quá","khá","cũng","vậy","thế","đây","đó","còn","vẫn",
}
EN_STOP = set(ENGLISH_STOP_WORDS)


def clean_text(text):
    \"\"\"Làm sạch text: xóa HTML, URL, ký tự đặc biệt, số.\"\"\"
    if not text or not isinstance(text, str):
        return ""
    text = re.sub(r'<[^>]+>',         ' ', text)   # xóa HTML tags
    text = re.sub(r'&[a-z]+;',        ' ', text)   # xóa HTML entities
    text = re.sub(r'https?://\\S+',    ' ', text)   # xóa URL
    text = re.sub(r'www\\.\\S+',       ' ', text)   # xóa www
    text = re.sub(r'\\S+@\\S+',        ' ', text)   # xóa địa chỉ email
    text = re.sub(r'[^\\w\\s]',        ' ', text)   # xóa ký tự đặc biệt
    text = re.sub(r'\\b\\d+\\b',        ' ', text)   # xóa số
    text = re.sub(r'\\s+',             ' ', text)   # chuẩn hóa khoảng trắng
    return text.lower().strip()


def detect_lang(text):
    \"\"\"Phát hiện ngôn ngữ chính (đơn giản, không cần thư viện).\"\"\"
    vn_chars = set('àáâãèéêìíòóôõùúýăđơưạặầẩẫậắẳẵặẹẽếềểễệỉịọộốồổỗợớờởỡụứừửữựỳỵ')
    count    = sum(1 for c in text if c in vn_chars)
    return 'vi' if count > len(text) * 0.04 else 'en'


def tokenize_vi(text):
    \"\"\"Tách từ tiếng Việt dùng underthesea.\"\"\"
    if UNDERTHESEA_OK:
        try:
            return vn_tokenize(text, format="text")
        except Exception:
            return text
    return text


def remove_stops(text):
    \"\"\"Xóa stopwords tiếng Việt và tiếng Anh, bỏ từ quá ngắn.\"\"\"
    return ' '.join(
        w for w in text.split()
        if len(w) >= 2 and w not in VN_STOP and w not in EN_STOP
    )


def preprocess(subject, body):
    \"\"\"Pipeline tiền xử lý đầy đủ cho 1 email.\"\"\"
    # Nhân tiêu đề 3 lần vì tiêu đề quan trọng hơn body
    raw = (str(subject) + ' ') * 3 + ' ' + str(body)
    cleaned = clean_text(raw)
    if detect_lang(cleaned) == 'vi':
        cleaned = tokenize_vi(cleaned)
    return remove_stops(cleaned)


print("✅ Hàm tiền xử lý sẵn sàng!")

# Ví dụ minh họa
sample_subj = "Khuyến mãi đặc biệt: giảm 50% hôm nay!"
sample_body = "Chúc mừng bạn đã trúng thưởng! Click vào link để nhận quà."
result = preprocess(sample_subj, sample_body)
print(f"\\n📝 Ví dụ tiền xử lý:")
print(f"   Input : {sample_subj}")
print(f"   Output: {result[:100]}")
"""))

# =============================================================================
# CELL 9 — ÁP DỤNG TIỀN XỬ LÝ
# =============================================================================
cells.append(code("""\
# ============================================================
# ÁP DỤNG TIỀN XỬ LÝ (CÓ CACHE)
# ============================================================

if os.path.exists(PROCESSED_CSV):
    print(f"📦 Tìm thấy cache: {PROCESSED_CSV}")
    df_proc = pd.read_csv(PROCESSED_CSV, encoding="utf-8-sig", low_memory=False)
    df_proc.fillna("", inplace=True)
    df_proc = df_proc.reset_index(drop=True)
    print(f"   ✅ Tải cache: {len(df_proc):,} email đã xử lý")
else:
    print("🔄 Đang tiền xử lý văn bản...")
    texts = []
    for _, row in tqdm(df.iterrows(), total=len(df), desc="⚙️  Tiền xử lý"):
        texts.append(preprocess(row.get("subject",""), row.get("body","")))
    df_proc = df.copy()
    df_proc["text"] = texts
    # Xóa email rỗng sau xử lý
    df_proc = df_proc[df_proc["text"].str.len() > 10].reset_index(drop=True)
    df_proc.to_csv(PROCESSED_CSV, index=False, encoding="utf-8-sig")
    print(f"   ✅ Đã lưu: {len(df_proc):,} email ({len(df)-len(df_proc)} đã lọc)")

# Hiển thị 3 mẫu trước/sau
print("\\n📝 Mẫu kết quả tiền xử lý (3 email):")
for i, (_, row) in enumerate(df_proc.sample(3, random_state=42).iterrows()):
    label_icon = "🔴" if row["label"] == SPAM else "🟢"
    print(f"\\n  [{i+1}] {label_icon} {row['label'].upper()}")
    print(f"       Subject : {str(row['subject'])[:60]}")
    print(f"       Text    : {str(row['text'])[:100]}...")
"""))

# =============================================================================
# CELL 10 — TF-IDF
# =============================================================================
cells.append(md("## 🔢 Phase 4 — Vector hóa TF-IDF\n\n**TF-IDF** (Term Frequency - Inverse Document Frequency) chuyển đổi văn bản thành vector số:\n- **TF**: Từ xuất hiện nhiều trong 1 email → trọng số cao\n- **IDF**: Từ hiếm trong toàn bộ dataset → trọng số cao\n- **TF-IDF = TF × IDF** → Từ đặc trưng của email đó nhận trọng số cao nhất"))

cells.append(code("""\
# ============================================================
# VECTOR HÓA TF-IDF
# ============================================================

X_text = df_proc["text"].fillna("").values   # numpy array
y_all  = df_proc["label"].values             # numpy array

print(f"🔢 Tạo TF-IDF vectorizer...")
print(f"   Max features : {MAX_FEATURES:,}")
print(f"   N-gram range : {NGRAM_RANGE}  (unigram + bigram)")
print(f"   Số email     : {len(X_text):,}")

tfidf = TfidfVectorizer(
    max_features = MAX_FEATURES,
    ngram_range  = NGRAM_RANGE,
    sublinear_tf = True,    # dùng log(1+tf) thay vì tf thô → giảm ảnh hưởng từ lặp nhiều
    min_df       = 2,       # bỏ từ chỉ xuất hiện trong < 2 email
    max_df       = 0.95,    # bỏ từ xuất hiện trong > 95% email (quá phổ biến)
)

X_tfidf = tfidf.fit_transform(X_text)

# Lưu vectorizer để dùng lại sau
with open(TFIDF_PKL, 'wb') as f:
    pickle.dump(tfidf, f)

print(f"\\n✅ Ma trận TF-IDF:")
print(f"   Kích thước  : {X_tfidf.shape[0]:,} email × {X_tfidf.shape[1]:,} features")
sparsity = 1 - X_tfidf.nnz / (X_tfidf.shape[0] * X_tfidf.shape[1])
print(f"   Độ thưa     : {sparsity*100:.1f}% (sparse matrix — tiết kiệm RAM)")
print(f"   💾 Đã lưu  : {TFIDF_PKL}")
"""))

# =============================================================================
# CELL 11 — CLASS IMBALANCE
# =============================================================================
cells.append(md("## ⚖️ Phase 5 — Xử lý Mất cân bằng Nhãn\n\n**Mất cân bằng nhãn** xảy ra khi Spam << Ham. Vấn đề: model sẽ \"lười\" — luôn đoán HAM để đạt accuracy cao mà không học được spam thật sự."))

cells.append(code("""\
# ============================================================
# CHIA DỮ LIỆU + XỬ LÝ MẤT CÂN BẰNG
# ============================================================

# QUAN TRỌNG: chia train/test TRƯỚC khi oversampling
# Lý do: oversampling chỉ áp dụng trên training set
#        Nếu oversample trước → test set bị "nhiễm" → kết quả đánh giá ảo
X_train, X_test, y_train, y_test, idx_train, idx_test = train_test_split(
    X_tfidf, y_all, np.arange(len(df_proc)),
    test_size=TEST_SIZE, random_state=RANDOM_STATE, stratify=y_all
)

print(f"📂 Chia dữ liệu train/test (80% / 20%):")
print(f"   Train: {X_train.shape[0]:,} email")
print(f"   Test : {X_test.shape[0]:,} email")

# Kiểm tra imbalance
spam_train = (y_train == SPAM).sum()
ham_train  = (y_train == HAM).sum()
ratio      = spam_train / (spam_train + ham_train)

print(f"\\n⚖️  Phân phối training set (trước oversampling):")
print(f"   Spam: {spam_train:,} ({ratio*100:.1f}%)")
print(f"   Ham : {ham_train:,}  ({(1-ratio)*100:.1f}%)")

if ratio < 0.30:
    print(f"\\n⚠️  Lệch nhãn nặng! → Áp dụng RandomOverSampler")
    print("   Kỹ thuật: Nhân bản ngẫu nhiên class thiểu số (Spam)")
    print("   Lý do chọn RandomOverSampler thay vì SMOTE:")
    print("   → SMOTE không hoạt động tốt với sparse TF-IDF matrix")

if IMBLEARN_OK:
    ros = RandomOverSampler(random_state=RANDOM_STATE)
    X_train_res, y_train_res = ros.fit_resample(X_train, y_train)
    spam_res = (y_train_res == SPAM).sum()
    ham_res  = (y_train_res == HAM).sum()
    print(f"\\n✅ Sau oversampling:")
    print(f"   Spam: {spam_res:,} | Ham: {ham_res:,}")

    # Biểu đồ trước/sau
    fig, axes = plt.subplots(1, 2, figsize=(12, 4))
    fig.suptitle("⚖️ Xử lý mất cân bằng nhãn (Training Set)",
                 fontsize=14, fontweight='bold')

    axes[0].bar(['SPAM','HAM'], [spam_train, ham_train],
                color=[SPAM_COLOR, HAM_COLOR], edgecolor='white')
    axes[0].set_title('TRƯỚC oversampling')
    axes[0].set_ylabel('Số lượng email')
    for i, v in enumerate([spam_train, ham_train]):
        axes[0].text(i, v + max(spam_train,ham_train)*0.01,
                     f'{v:,}', ha='center', fontweight='bold')

    axes[1].bar(['SPAM','HAM'], [spam_res, ham_res],
                color=[SPAM_COLOR, HAM_COLOR], edgecolor='white')
    axes[1].set_title('SAU oversampling')
    axes[1].set_ylabel('Số lượng email')
    for i, v in enumerate([spam_res, ham_res]):
        axes[1].text(i, v + max(spam_res,ham_res)*0.01,
                     f'{v:,}', ha='center', fontweight='bold')

    plt.tight_layout()
    plt.savefig(os.path.join(REPORTS_DIR, 'class_balance.png'),
                dpi=150, bbox_inches='tight', facecolor='white')
    plt.show()
else:
    X_train_res, y_train_res = X_train, y_train
    print("⚠️  Bỏ qua oversampling (imbalanced-learn chưa cài)")
"""))

# =============================================================================
# CELL 12 — TRAIN MODEL
# =============================================================================
cells.append(md("## 🤖 Phase 6 — Huấn luyện Mô hình Naive Bayes\n\n**Naive Bayes hoạt động như thế nào?**\n\nCông thức Bayes:\n$$P(spam \\mid email) \\propto P(spam) \\times \\prod_{i} P(từ_i \\mid spam)$$\n\n- Với mỗi email, tính xác suất là spam dựa trên xác suất của từng từ trong email đó\n- \"Naive\" = giả định các từ độc lập nhau (đơn giản hóa nhưng thực tế vẫn rất hiệu quả)\n- **Laplace smoothing** (alpha=1): tránh xác suất = 0 cho từ chưa gặp trong training"))

cells.append(code("""\
# ============================================================
# HUẤN LUYỆN MULTINOMIAL NAIVE BAYES
# ============================================================

print(f"🤖 Đang huấn luyện Multinomial Naive Bayes...")
print(f"   Training: {X_train_res.shape[0]:,} email")
print(f"   Alpha (Laplace smoothing): 1.0")

model = MultinomialNB(alpha=1.0)
model.fit(X_train_res, y_train_res)

# Lưu model để dùng lại
with open(MODEL_PKL, 'wb') as f:
    pickle.dump(model, f)

# Dự đoán trên test set
y_pred = model.predict(X_test)

print(f"\\n✅ Huấn luyện hoàn tất!")
print(f"   💾 Model đã lưu: {MODEL_PKL}")
print(f"   📊 Accuracy sơ bộ: {accuracy_score(y_test, y_pred)*100:.2f}%")
"""))

# =============================================================================
# CELL 13 — ĐÁNH GIÁ
# =============================================================================
cells.append(md("## 📈 Phase 7 — Đánh giá Mô hình"))

cells.append(code("""\
# ============================================================
# TÍNH ĐẦY ĐỦ CÁC CHỈ SỐ ĐÁNH GIÁ
# ============================================================

acc  = accuracy_score(y_test, y_pred)
prec = precision_score(y_test, y_pred, pos_label=SPAM, zero_division=0)
rec  = recall_score(y_test, y_pred,    pos_label=SPAM, zero_division=0)
f1   = f1_score(y_test, y_pred,        pos_label=SPAM, zero_division=0)

metrics_data = [
    ["Accuracy",  f"{acc*100:.2f}%",
     "Tổng số email dự đoán đúng / tổng email"],
    ["Precision", f"{prec*100:.2f}%",
     "Trong các email bị đoán là SPAM, bao nhiêu % thực sự là SPAM?"],
    ["Recall",    f"{rec*100:.2f}%",
     "Trong tất cả SPAM thật, bao nhiêu % được phát hiện?"],
    ["F1-Score",  f"{f1*100:.2f}%",
     "Trung bình điều hòa của Precision và Recall"],
]
print("📈 KẾT QUẢ ĐÁNH GIÁ MÔ HÌNH:")
print(tabulate(metrics_data, headers=["Chỉ số","Giá trị","Ý nghĩa"], tablefmt="grid"))

print("\\n📋 Báo cáo phân loại đầy đủ:")
print(classification_report(y_test, y_pred, target_names=['Ham','Spam']))

# Phân tích kết quả
print("\\n🔍 PHÂN TÍCH:")
if acc >= 0.95:
    print(f"  ✅ Accuracy {acc*100:.1f}% → Mô hình ĐẠT YÊU CẦU (>= 95%)")
elif acc >= 0.90:
    print(f"  🟡 Accuracy {acc*100:.1f}% → Mô hình TỐT nhưng có thể cải thiện")
else:
    print(f"  ⚠️  Accuracy {acc*100:.1f}% → Mô hình CẦN CẢI THIỆN")

if rec < 0.70:
    print(f"  ⚠️  Recall {rec*100:.1f}% thấp → Nhiều spam bị bỏ sót")
    print("      Gợi ý: Giảm threshold, thêm spam features, tăng training data")
else:
    print(f"  ✅ Recall {rec*100:.1f}% → Phát hiện spam tốt")
"""))

# =============================================================================
# CELL 14 — CONFUSION MATRIX + BAR CHART
# =============================================================================
cells.append(code("""\
# ============================================================
# CONFUSION MATRIX + BIỂU ĐỒ ĐÁNH GIÁ
# ============================================================

cm = confusion_matrix(y_test, y_pred, labels=[HAM, SPAM])
tn, fp, fn, tp = cm.ravel()

fig, axes = plt.subplots(1, 2, figsize=(14, 5))
fig.suptitle("📈 Kết quả đánh giá — Multinomial Naive Bayes",
             fontsize=14, fontweight='bold')

# --- Confusion Matrix heatmap ---
labels_cm = [['TN\\n(Ham đúng)', 'FP\\n(Ham → Spam nhầm)'],
             ['FN\\n(Spam bỏ sót)', 'TP\\n(Spam đúng)']]
annot_cm  = np.array([[f'{tn:,}\\n{tn/(tn+fp)*100:.1f}%',
                        f'{fp:,}\\n{fp/(tn+fp)*100:.1f}%'],
                       [f'{fn:,}\\n{fn/(fn+tp)*100:.1f}%',
                        f'{tp:,}\\n{tp/(fn+tp)*100:.1f}%']])
sns.heatmap(cm, annot=annot_cm, fmt='', cmap='Blues',
            xticklabels=['Đoán HAM', 'Đoán SPAM'],
            yticklabels=['Thực tế HAM', 'Thực tế SPAM'],
            ax=axes[0], linewidths=1, linecolor='white',
            annot_kws={"size": 11, "weight": "bold"})
axes[0].set_title('Confusion Matrix', fontweight='bold')
axes[0].set_ylabel('Nhãn thực tế')
axes[0].set_xlabel('Nhãn dự đoán')

# --- Bar chart các chỉ số ---
m_names  = ['Accuracy', 'Precision', 'Recall', 'F1-Score']
m_values = [acc, prec, rec, f1]
bar_cols  = ['#2ecc71' if v >= 0.90 else '#f39c12' if v >= 0.80 else '#e74c3c'
             for v in m_values]
bars = axes[1].bar(m_names, [v*100 for v in m_values],
                   color=bar_cols, edgecolor='white', linewidth=1.5, width=0.5)
axes[1].set_title('Các chỉ số đánh giá', fontweight='bold')
axes[1].set_ylabel('Giá trị (%)')
axes[1].set_ylim(0, 108)
axes[1].axhline(y=90, color='green', linestyle='--', alpha=0.5, label='Ngưỡng 90%')
axes[1].legend()
for bar, val in zip(bars, m_values):
    axes[1].text(bar.get_x() + bar.get_width()/2,
                 bar.get_height() + 0.8,
                 f'{val*100:.1f}%', ha='center', fontweight='bold')
axes[1].grid(axis='y', alpha=0.3)

plt.tight_layout()
plt.savefig(os.path.join(REPORTS_DIR, 'evaluation.png'),
            dpi=150, bbox_inches='tight', facecolor='white')
plt.show()

print(f"   TN={tn:,}  FP={fp:,}  FN={fn:,}  TP={tp:,}")
"""))

# =============================================================================
# CELL 15 — FEATURE IMPORTANCE
# =============================================================================
cells.append(code("""\
# ============================================================
# TOP TỪ QUAN TRỌNG NHẤT TRONG PHÂN LOẠI
# ============================================================

feat_names = np.array(tfidf.get_feature_names_out())

# Chỉ số class spam vs ham
spam_idx = list(model.classes_).index(SPAM)
ham_idx  = list(model.classes_).index(HAM)

# Điểm = log_prob_spam - log_prob_ham
# Dương cao → từ đặc trưng của spam
# Âm lớn  → từ đặc trưng của ham
scores = model.feature_log_prob_[spam_idx] - model.feature_log_prob_[ham_idx]

top_n = 20
top_spam_idx = np.argsort(scores)[-top_n:][::-1]
top_ham_idx  = np.argsort(scores)[:top_n]

top_spam_words  = feat_names[top_spam_idx]
top_spam_scores = scores[top_spam_idx]
top_ham_words   = feat_names[top_ham_idx]
top_ham_scores  = -scores[top_ham_idx]

fig, axes = plt.subplots(1, 2, figsize=(16, 7))
fig.suptitle(f"🔑 Top {top_n} từ đặc trưng nhất cho phân loại",
             fontsize=14, fontweight='bold')

# Top spam words
y_pos = np.arange(top_n)
axes[0].barh(y_pos, top_spam_scores[::-1], color=SPAM_COLOR, alpha=0.8)
axes[0].set_yticks(y_pos)
axes[0].set_yticklabels(top_spam_words[::-1], fontsize=9)
axes[0].set_title(f'Top {top_n} từ đặc trưng SPAM',
                  fontweight='bold', color=SPAM_COLOR)
axes[0].set_xlabel('Score (spam - ham log-prob)')
axes[0].grid(axis='x', alpha=0.3)

# Top ham words
axes[1].barh(y_pos, top_ham_scores[::-1], color=HAM_COLOR, alpha=0.8)
axes[1].set_yticks(y_pos)
axes[1].set_yticklabels(top_ham_words[::-1], fontsize=9)
axes[1].set_title(f'Top {top_n} từ đặc trưng HAM',
                  fontweight='bold', color=HAM_COLOR)
axes[1].set_xlabel('Score (ham - spam log-prob)')
axes[1].grid(axis='x', alpha=0.3)

plt.tight_layout()
plt.savefig(os.path.join(REPORTS_DIR, 'feature_importance.png'),
            dpi=150, bbox_inches='tight', facecolor='white')
plt.show()

# Bảng top 10
print("🔴 TOP 10 TỪ ĐẶC TRƯNG SPAM:")
print(tabulate([[w, f"{s:.3f}"] for w, s in zip(top_spam_words[:10], top_spam_scores[:10])],
               headers=["Từ", "Score"], tablefmt="grid"))
print("\\n🟢 TOP 10 TỪ ĐẶC TRƯNG HAM:")
print(tabulate([[w, f"{s:.3f}"] for w, s in zip(top_ham_words[:10], top_ham_scores[:10])],
               headers=["Từ", "Score"], tablefmt="grid"))
"""))

# =============================================================================
# CELL 16 — XUẤT BÁO CÁO EXCEL
# =============================================================================
cells.append(md("## 📋 Phase 8 — Xuất Báo cáo"))

cells.append(code("""\
# ============================================================
# XUẤT BÁO CÁO EXCEL (4 SHEET)
# ============================================================

wb = openpyxl.Workbook()

# Style chung
H_FONT  = Font(bold=True, color="FFFFFF", size=11)
H_FILL1 = PatternFill("solid", fgColor="2C3E50")  # tiêu đề
H_FILL2 = PatternFill("solid", fgColor="2980B9")  # sub-header
CENTER  = Alignment(horizontal="center", vertical="center")
SPAM_FILL = PatternFill("solid", fgColor="FADBD8")
HAM_FILL  = PatternFill("solid", fgColor="D5F5E3")

def style_header(ws, row_num=1, fill=None):
    fill = fill or H_FILL1
    for cell in ws[row_num]:
        cell.font      = H_FONT
        cell.fill      = fill
        cell.alignment = CENTER

def autofit(ws):
    for col in ws.columns:
        max_w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 3, 55)

# ---- SHEET 1: Thống kê dataset ----
ws1 = wb.active
ws1.title = "1.Thong ke"
ws1.append(["CHỈ SỐ", "GIÁ TRỊ", "GHI CHÚ"])
style_header(ws1)
g1_n = int((df_proc.get('source', pd.Series()) == 'gmail1').sum()) if 'source' in df_proc.columns else 0
g2_n = int((df_proc.get('source', pd.Series()) == 'gmail2').sum()) if 'source' in df_proc.columns else 0
ws1_data = [
    ["Tổng số email",            len(df_proc),                   ""],
    ["Email SPAM",                int(spam_n),                    f"{spam_n/len(df_proc)*100:.1f}%"],
    ["Email HAM",                 int(ham_n),                     f"{ham_n/len(df_proc)*100:.1f}%"],
    ["Gmail 1 (wiisch3009)",      g1_n,                           "Nhãn heuristics"],
    ["Gmail 2 (nguyenkhanhtoan)", g2_n,                           "Nhãn thật từ Gmail"],
    ["Train set",                 X_train_res.shape[0],           f"{(1-TEST_SIZE)*100:.0f}%"],
    ["Test set",                  X_test.shape[0],                f"{TEST_SIZE*100:.0f}%"],
    ["TF-IDF features",           X_tfidf.shape[1],               f"max={MAX_FEATURES:,}"],
    ["Mô hình",                   "Multinomial Naive Bayes",      "alpha=1.0"],
]
for row in ws1_data:
    ws1.append(row)
autofit(ws1)

# ---- SHEET 2: Kết quả mô hình ----
ws2 = wb.create_sheet("2.Ket qua mo hinh")
ws2.append(["CHỈ SỐ", "GIÁ TRỊ", "Ý NGHĨA"])
style_header(ws2)
ws2_data = [
    ["Accuracy",        f"{acc*100:.4f}%",  "Tỉ lệ dự đoán đúng tổng thể"],
    ["Precision",       f"{prec*100:.4f}%", "Khi đoán spam, bao nhiêu % đúng"],
    ["Recall",          f"{rec*100:.4f}%",  "Bao nhiêu % spam thật được phát hiện"],
    ["F1-Score",        f"{f1*100:.4f}%",   "Cân bằng giữa Precision và Recall"],
    ["True Positive",   int(tp),            "Spam được phát hiện đúng"],
    ["True Negative",   int(tn),            "Ham được phân loại đúng"],
    ["False Positive",  int(fp),            "Ham bị nhầm là spam (Type I error)"],
    ["False Negative",  int(fn),            "Spam bị bỏ sót (Type II error)"],
]
for row in ws2_data:
    ws2.append(row)
autofit(ws2)

# ---- SHEET 3: Email + nhãn dự đoán ----
ws3 = wb.create_sheet("3.Email+Nhan du doan")
ws3.append(["STT","TIÊU ĐỀ","NGƯỜI GỬI","NHÃN THẬT","NHÃN DỰ ĐOÁN","ĐÚNG/SAI"])
style_header(ws3)

test_df = df_proc.iloc[idx_test].copy().reset_index(drop=True)
test_df["predicted"] = y_pred

for i, row in test_df.head(2000).iterrows():
    ok   = "✓" if row["label"] == row["predicted"] else "✗"
    ws3.append([
        i+1,
        str(row.get("subject",""))[:70],
        str(row.get("sender", ""))[:50],
        row["label"],
        row["predicted"],
        ok
    ])
    # Tô màu theo nhãn dự đoán
    fill = SPAM_FILL if row["predicted"] == SPAM else HAM_FILL
    for col_idx in range(1, 7):
        ws3.cell(row=i+2, column=col_idx).fill = fill
autofit(ws3)

# ---- SHEET 4: Email dự đoán sai ----
ws4 = wb.create_sheet("4.Email du doan sai")
ws4.append(["STT","TIÊU ĐỀ","NGƯỜI GỬI","NHÃN THẬT","NHÃN DỰ ĐOÁN","PHÂN TÍCH"])
style_header(ws4, fill=PatternFill("solid", fgColor="C0392B"))

wrong_df = test_df[test_df["label"] != test_df["predicted"]].reset_index(drop=True)
for i, row in wrong_df.head(500).iterrows():
    analysis = (
        "Spam bị bỏ sót (False Negative)" if row["label"]==SPAM
        else "Ham bị nhầm là Spam (False Positive)"
    )
    ws4.append([
        i+1,
        str(row.get("subject",""))[:70],
        str(row.get("sender", ""))[:50],
        row["label"],
        row["predicted"],
        analysis
    ])
autofit(ws4)

wb.save(REPORT_XLSX)
print(f"✅ Đã xuất báo cáo Excel: {REPORT_XLSX}")
print(f"   📄 Sheet 1: Thống kê dataset")
print(f"   📄 Sheet 2: Kết quả mô hình ({len(ws2_data)} chỉ số)")
print(f"   📄 Sheet 3: {len(test_df.head(2000))} email + nhãn dự đoán")
print(f"   📄 Sheet 4: {len(wrong_df.head(500))} email dự đoán sai")
"""))

# =============================================================================
# CELL 17 — TÓM TẮT CUỐI
# =============================================================================
cells.append(code("""\
# ============================================================
# TÓM TẮT TOÀN BỘ DỰ ÁN
# ============================================================

summary = f\"\"\"
╔══════════════════════════════════════════════════════════════╗
║      BÁO CÁO CUỐI — SPAM EMAIL CLASSIFIER                  ║
║      Chủ đề 16 — UTH AI Project                            ║
╠══════════════════════════════════════════════════════════════╣
║  📊 DỮ LIỆU                                                ║
║    Tổng email      : {len(df_proc):>7,}                          ║
║    Spam            : {int(spam_n):>7,}  ({spam_n/len(df_proc)*100:5.1f}%)              ║
║    Ham             : {int(ham_n):>7,}  ({ham_n/len(df_proc)*100:5.1f}%)              ║
║    Nguồn           : 2 Gmail (mbox + CSV)                  ║
╠══════════════════════════════════════════════════════════════╣
║  🤖 MÔ HÌNH                                                ║
║    Thuật toán      : Multinomial Naive Bayes               ║
║    Tỉ lệ train/test: {int((1-TEST_SIZE)*100)}% / {int(TEST_SIZE*100)}%                          ║
║    Oversampling    : RandomOverSampler                     ║
║    TF-IDF features : {X_tfidf.shape[1]:>7,} features                   ║
╠══════════════════════════════════════════════════════════════╣
║  📈 KẾT QUẢ ĐÁNH GIÁ                                       ║
║    Accuracy        : {acc*100:>7.2f}%                            ║
║    Precision       : {prec*100:>7.2f}%                            ║
║    Recall          : {rec*100:>7.2f}%                            ║
║    F1-Score        : {f1*100:>7.2f}%                            ║
║    TP={tp:,}  TN={tn:,}  FP={fp:,}  FN={fn:,}               ║
╠══════════════════════════════════════════════════════════════╣
║  📁 FILE ĐẦU RA                                            ║
║    output/labeled_dataset.csv    (dữ liệu đã gán nhãn)    ║
║    output/processed_dataset.csv  (dữ liệu đã xử lý)       ║
║    output/models/naive_bayes_model.pkl                     ║
║    output/models/tfidf_vectorizer.pkl                      ║
║    output/reports/spam_report.xlsx (4 sheets)              ║
║    output/reports/eda_analysis.png                         ║
║    output/reports/class_balance.png                        ║
║    output/reports/evaluation.png                           ║
║    output/reports/feature_importance.png                   ║
╚══════════════════════════════════════════════════════════════╝
\"\"\"
print(summary)

# Lưu file text
with open(REPORT_TXT, 'w', encoding='utf-8') as f:
    f.write(summary)
print(f"✅ Đã lưu tóm tắt: {REPORT_TXT}")

# Hiển thị bảng biểu đồ tổng kết
print("\\n📊 TẤT CẢ BIỂU ĐỒ ĐÃ LƯU:")
img_files = [f for f in os.listdir(REPORTS_DIR) if f.endswith('.png')]
for img in sorted(img_files):
    print(f"   📈 {img}")
"""))

# =============================================================================
# CELL CUỐI — KẾT LUẬN
# =============================================================================
cells.append(md("""\
## 🎯 Kết luận

### Những gì đã thực hiện
1. **Thu thập dữ liệu** — 2 tài khoản Gmail (~39,000 email thực)
2. **Gán nhãn** — Nhãn thật từ Gmail (Gmail 2) + Heuristics (Gmail 1)
3. **EDA** — Phân tích phân phối, độ dài, sender domain
4. **Tiền xử lý** — Làm sạch HTML/URL, tách từ tiếng Việt (underthesea), xóa stopwords
5. **TF-IDF** — Vector hóa với n-gram (1,2), 50,000 features
6. **Xử lý imbalance** — RandomOverSampler trên training set
7. **Naive Bayes** — Multinomial NB với Laplace smoothing
8. **Đánh giá** — Accuracy, Precision, Recall, F1, Confusion Matrix
9. **Báo cáo** — Excel 4 sheets + ảnh biểu đồ + summary text

### Hướng cải thiện
- Thêm dữ liệu spam đa dạng hơn (hiện tại chỉ ~12%)
- Thử nghiệm các mô hình khác: SVM, Random Forest, BERT
- Cải thiện pipeline tiếng Việt với PhoBERT
- Triển khai thành web app với FastAPI
"""))

# ============================================================
# TẠO FILE NOTEBOOK
# ============================================================
nb.cells = cells

output_path = os.path.join(os.path.dirname(__file__), "spam_classifier.ipynb")
with open(output_path, 'w', encoding='utf-8') as f:
    nbf.write(nb, f)

print(f"✅ Đã tạo: {output_path}")
print(f"   Số cells: {len(cells)}")
print(f"   Mở trong VSCode: File → Open → spam_classifier.ipynb")
